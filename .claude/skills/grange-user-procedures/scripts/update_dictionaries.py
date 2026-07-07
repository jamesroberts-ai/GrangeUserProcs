"""Regenerate the skill's dictionary CSVs from freshly downloaded Grange
dictionary workbooks (e.g. Dictionary - Winemaking.xlsm, Grange Dictionaries.xlsx).

Usage:
    python3 update_dictionaries.py <workbook1> [workbook2 ...] --out <dictionaries dir>

Expected sheet layout (as produced by the Grange dictionary download macros):
  Row 1: 'File name = ' | <GRANGE FILE NAME> | ... | 'Downloaded' | <date(s)>
  Row 3: headers (Field Name, Type, Field No, Field Definition, Conversion Code,
         Column Heading, Output Format, Depth, Assoc, DESC, Key, Examples, Comment)
  Row 4+: field definitions

Dedupe rule: where the same Grange file appears in more than one workbook/sheet,
the version with the latest 'Downloaded' date wins.

After running, manually check for near-duplicate file names caused by typos in
the workbooks (e.g. 'BLEND.PROP AMD ...' vs 'BLEND.PROP AND ...') and merge,
and keep INDEX.md's file list in sync if new Grange files appear.
"""
import argparse, csv, datetime, os, re, sys
from openpyxl import load_workbook

HEADERS = ["Field Name", "Type", "Field No", "Field Definition", "Conversion Code",
           "Column Heading", "Output Format", "Depth", "Assoc", "DESC", "Key",
           "Examples", "Comment"]

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbooks", nargs="+")
    ap.add_argument("--out", required=True, help="dictionaries output directory")
    args = ap.parse_args()
    os.makedirs(args.out, exist_ok=True)

    catalog = {}  # FILE NAME -> (download_date, source_label, data rows)
    for path in args.workbooks:
        label = os.path.basename(path)
        wb = load_workbook(path, read_only=True)
        for sheet in wb.sheetnames:
            rows = list(wb[sheet].iter_rows(values_only=True))
            if not rows or not rows[0] or rows[0][0] != "File name = ":
                continue
            fname = str(rows[0][1]).strip()
            dates = [c for c in rows[0] if isinstance(c, datetime.datetime)]
            dl = max(dates) if dates else datetime.datetime(1900, 1, 1)
            data = []
            for r in rows[3:]:
                if r is None:
                    continue
                vals = list(r[:13]) + [None] * (13 - len(r[:13]))
                if all(v in (None, "") for v in vals):
                    continue
                if vals[0] is None or str(vals[0]).startswith("="):
                    continue
                data.append(["" if v is None else str(v).strip() for v in vals])
            if not data:
                continue
            key = fname.upper()
            if key not in catalog or dl > catalog[key][0]:
                catalog[key] = (dl, f"{label}:{sheet}", data)
        wb.close()

    for fname, (dl, src, data) in sorted(catalog.items()):
        safe = re.sub(r"[^A-Za-z0-9._-]+", "_", fname)
        with open(os.path.join(args.out, f"{safe}.csv"), "w", newline="") as f:
            w = csv.writer(f)
            w.writerow([f"# Grange file: {fname}", f"source: {src}",
                        f"downloaded: {dl.date()}"])
            w.writerow(HEADERS)
            w.writerows(data)
        print(f"{fname:45s} {len(data):5d} fields   {dl.date()}  <- {src}")
    print(f"\n{len(catalog)} dictionaries written to {args.out}", file=sys.stderr)

if __name__ == "__main__":
    main()
